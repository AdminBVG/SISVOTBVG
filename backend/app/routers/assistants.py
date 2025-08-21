from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import csv
from io import StringIO, BytesIO
from pathlib import Path
from openpyxl import load_workbook, Workbook

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import Response

from .. import models, schemas, database
from ..security import get_current_user, require_role

router = APIRouter(prefix="/elections/{election_id}/assistants", tags=["assistants"])


def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/template", dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])] )
def export_template(election_id: int, format: str = "csv"):
    headers = ["id", "accionista", "representante_legal", "apoderado", "acciones"]
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        bio = BytesIO()
        wb.save(bio)
        return Response(
            content=bio.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=padron_template.xlsx"},
        )
    else:
        content = ",".join(headers) + "\n"
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=padron_template.csv"},
        )

@router.post(
    "/import-excel",
    response_model=List[schemas.Attendee],
    dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])]
)
def import_attendees_excel(
    election_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    required = {"id", "accionista", "representante_legal", "apoderado", "acciones"}
    results: List[models.Attendee] = []
    errors: List[str] = []
    seen_ids = set(
        a.identifier
        for a in db.query(models.Attendee.identifier).filter_by(election_id=election_id)
    )

    def process_row(row: dict, idx: int):
        row_errors = []
        identifier = str(row.get("id", "")).strip()
        if not identifier:
            row_errors.append("id required")
        elif identifier in seen_ids:
            row_errors.append("duplicate id")
        accionista = (row.get("accionista") or "").strip()
        if not accionista:
            row_errors.append("accionista required")
        acciones_raw = row.get("acciones")
        try:
            acciones_val = float(acciones_raw)
            if acciones_val <= 0:
                row_errors.append("acciones must be positive")
        except Exception:
            row_errors.append("acciones must be numeric")

        representante = (row.get("representante_legal") or None)
        apoderado = (row.get("apoderado") or None)

        if row_errors:
            errors.append(f"Row {idx}: {', '.join(row_errors)}")
            return

        attendee = models.Attendee(
            election_id=election_id,
            identifier=identifier,
            accionista=accionista,
            representante=representante,
            apoderado=apoderado,
            acciones=acciones_val,
        )
        db.add(attendee)
        results.append(attendee)
        seen_ids.add(identifier)

        # Sync with shareholders/attendance so registrars can see attendees
        sh = (
            db.query(models.Shareholder)
            .filter_by(code=attendee.identifier)
            .first()
        )
        if not sh:
            sh = models.Shareholder(
                code=attendee.identifier,
                name=attendee.accionista,
                document=attendee.identifier,
                actions=attendee.acciones,
                status="ACTIVE",
            )
            db.add(sh)
            db.flush()
        else:
            sh.name = attendee.accionista
            sh.actions = attendee.acciones
            db.flush()
        if not db.query(models.Attendance).filter_by(
            election_id=election_id, shareholder_id=sh.id
        ).first():
            db.add(
                models.Attendance(
                    election_id=election_id,
                    shareholder_id=sh.id,
                    mode=models.AttendanceMode.AUSENTE,
                    present=False,
                )
            )

    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(file.file)
        sheet = wb.active
        rows = list(sheet.values)
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        if not required.issubset(headers):
            missing = required - set(headers)
            raise HTTPException(
                status_code=400,
                detail=f"Missing columns: {', '.join(sorted(missing))}",
            )
        for idx, data_row in enumerate(rows[1:], start=2):
            row = dict(zip(headers, data_row))
            process_row(row, idx)
    else:
        content = file.file.read().decode("utf-8")
        reader = csv.DictReader(StringIO(content))
        if not required.issubset(reader.fieldnames or []):
            missing = required - set(reader.fieldnames or [])
            raise HTTPException(
                status_code=400, detail=f"Missing columns: {', '.join(sorted(missing))}"
            )
        for idx, row in enumerate(reader, start=2):
            process_row(row, idx)

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    db.commit()
    output: List[schemas.Attendee] = []
    for att in results:
        db.refresh(att)
        data = schemas.Attendee.model_validate(att).model_dump()
        data["requires_document"] = bool(att.apoderado)
        data["document_uploaded"] = bool(att.apoderado_pdf_url)
        output.append(schemas.Attendee(**data))
    return output


@router.get(
    "",
    response_model=List[schemas.Attendee],
    dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])]
)
def list_attendees(election_id: int, db: Session = Depends(get_db)):
    attendees = db.query(models.Attendee).filter_by(election_id=election_id).all()
    result: List[schemas.Attendee] = []
    for att in attendees:
        data = schemas.Attendee.model_validate(att).model_dump()
        data["requires_document"] = bool(att.apoderado)
        data["document_uploaded"] = bool(att.apoderado_pdf_url)
        result.append(schemas.Attendee(**data))
    return result


@router.get(
    "/{attendee_id}",
    response_model=schemas.Attendee,
    dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])]
)
def get_attendee(
    election_id: int,
    attendee_id: int,
    db: Session = Depends(get_db),
):
    attendee = db.query(models.Attendee).filter_by(id=attendee_id, election_id=election_id).first()
    if not attendee:
        raise HTTPException(status_code=404, detail="attendee not found")
    data = schemas.Attendee.model_validate(attendee).model_dump()
    data["requires_document"] = bool(attendee.apoderado)
    data["document_uploaded"] = bool(attendee.apoderado_pdf_url)
    return schemas.Attendee(**data)


@router.put(
    "/{attendee_id}",
    response_model=schemas.Attendee,
    dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])]
)
def update_attendee(
    election_id: int,
    attendee_id: int,
    payload: schemas.AttendeeUpdate,
    db: Session = Depends(get_db),
):
    attendee = db.query(models.Attendee).filter_by(id=attendee_id, election_id=election_id).first()
    if not attendee:
        raise HTTPException(status_code=404, detail="attendee not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field == "apoderado" and not value:
            attendee.apoderado_pdf_url = None
        setattr(attendee, field, value)
    db.commit()
    db.refresh(attendee)
    data = schemas.Attendee.model_validate(attendee).model_dump()
    data["requires_document"] = bool(attendee.apoderado)
    data["document_uploaded"] = bool(attendee.apoderado_pdf_url)
    return schemas.Attendee(**data)


@router.delete(
    "/{attendee_id}",
    status_code=204,
    dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])]
)
def delete_attendee(
    election_id: int,
    attendee_id: int,
    db: Session = Depends(get_db),
):
    attendee = db.query(models.Attendee).filter_by(id=attendee_id, election_id=election_id).first()
    if not attendee:
        raise HTTPException(status_code=404, detail="attendee not found")
    db.delete(attendee)
    db.commit()


@router.post(
    "/{attendee_id}/apoderado-pdf",
    response_model=schemas.Attendee,
    dependencies=[require_role(["FUNCIONAL_BVG", "ADMIN_BVG"])]
)
async def upload_apoderado_pdf(
    election_id: int,
    attendee_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    attendee = db.query(models.Attendee).filter_by(id=attendee_id, election_id=election_id).first()
    if not attendee:
        raise HTTPException(status_code=404, detail="attendee not found")
    if not attendee.apoderado:
        raise HTTPException(status_code=400, detail="attendee has no apoderado")
    if attendee.apoderado_pdf_url:
        raise HTTPException(status_code=400, detail="document already uploaded")
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="invalid file type")
    content = await file.read()
    storage_dir = Path("storage") / str(election_id) / "apoderados"
    storage_dir.mkdir(parents=True, exist_ok=True)
    file_path = storage_dir / f"{attendee.id}.pdf"
    with open(file_path, "wb") as f:
        f.write(content)
    attendee.apoderado_pdf_url = str(file_path)
    db.commit()
    db.refresh(attendee)
    data = schemas.Attendee.model_validate(attendee).model_dump()
    data["requires_document"] = True
    data["document_uploaded"] = True
    return schemas.Attendee(**data)
